"""Reader for the BDF "Listungen" input sheet.

The Amazon template is stable. The input sheet is not: across the three samples
we have, the header sits on row 1 or row 5, the column count is 50/51/53, and
names drift ("GTIN with Zeroes" / "GTIN mit Zeroes", "Inhalt" /
"Einheit Netto-füll-menge"). So nothing here is positional. We find the header row
by looking for anchor columns, then match every column by normalised name against
a synonym table.

A source column we need but cannot find is a hard error, never a silent skip.
"""

from __future__ import annotations

import re
import unicodedata
import warnings
from pathlib import Path

from openpyxl import load_workbook

warnings.filterwarnings("ignore")

# Columns that must be present for the sheet to be a Listungen sheet at all.
HEADER_ANCHORS = ("nart", "artikellangtext1")

# canonical name -> accepted header spellings
SYNONYMS = {
    "nart": ["NART"],
    "title": ["Artikellangtext 1", "Artikellangtext1"],
    "gtin": ["GTIN Stück", "GTIN Stueck", "GTIN Stück2"],
    "gtin_padded": ["GTIN mit Zeroes", "GTIN with Zeroes", "GTIN mit Zeros"],
    "mgr_name": ["MGR Bezeichnung"],
    "mgr": ["MGR"],
    "template_hint": ["Template"],
    "launch_date": ["KWLT"],
    "net_content": ["Netto-füll-menge", "Nettofüllmenge"],
    "net_content_unit": ["Einheit Netto-füll-menge", "Inhalt"],
    "origin": ["Her-kunfts-land", "Herkunftsland"],
    "width": ["Breite Einzelstück"],
    "height": ["Höhe Einzelstück"],
    "length": ["Länge/Tiefe Einzelstück", "Länge Einzelstück"],
    "case_width": ["Breite VE"],
    "case_height": ["Höhe VE"],
    "case_length": ["Länge VE"],
    "case_weight": ["Gewicht VE KG"],
    "units_per_case": ["Stück pro VE"],
    "weight_gross": ["Gewicht Einzelstück (brutto) KG"],
    "weight_net": ["Gewicht Einzelstück Kg (netto)"],
    "dg_profile": ["Gefahrgut-profil", "Gefahrgutprofil"],
    "un_number": ["UN Nummer", "UN-Nummer"],
    "cost_price": ["Cost Price Amazon"],
    "uvp": ["UVP*", "UVP"],
    "ingredients": ["Incis?", "Incis", "INCI"],
    "short_description": ["short description", "Kurzbeschreibung"],
    "long_description": ["long description", "Langbeschreibung"],
    "gender": ["Zielgruppe (Geschlecht/Gender)", "Zielgruppe"],
    "skin_type": ["Hauttyp(normal, sensitive)", "Hauttyp (normal, sensitive)", "Hauttyp"],
    "scent": ["Duft (Scent)", "Duft"],
    "sales_status": ["Verkaufsstatus"],
    "predecessor_asin": ["ASIN"],
}

REQUIRED_CANONICAL = ["nart", "title", "gtin", "launch_date", "net_content", "origin"]


def normalise(s) -> str:
    if s is None:
        return ""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.replace("ß", "ss")
    return re.sub(r"[^a-z0-9]", "", s.lower())


_SYNONYM_INDEX = {}
for canon, spellings in SYNONYMS.items():
    for sp in spellings:
        _SYNONYM_INDEX[normalise(sp)] = canon


class InputSheetError(Exception):
    pass


class ListungenSheet:
    def __init__(self, path, sheet=None):
        self.path = Path(path)
        wb = load_workbook(self.path, data_only=False)
        self.wb_values = load_workbook(self.path, data_only=True)

        self.sheet_name = sheet or self._pick_sheet(wb)
        self.ws = wb[self.sheet_name]
        self.ws_values = self.wb_values[self.sheet_name]

        self.header_row = self._find_header_row(self.ws)
        self.columns, self.unmapped = self._map_columns(self.ws, self.header_row)

        missing = [c for c in REQUIRED_CANONICAL if c not in self.columns]
        if missing:
            raise InputSheetError(
                f"{self.path.name}: input sheet is missing required column(s): "
                + ", ".join(missing)
                + f". Detected header row {self.header_row}, "
                + f"mapped {len(self.columns)} columns."
            )

    # ------------------------------------------------------------------ setup

    def _pick_sheet(self, wb):
        best, best_score = None, -1
        for ws in wb.worksheets:
            score = 0
            for r in range(1, min(ws.max_row, 12) + 1):
                hits = sum(
                    1
                    for c in range(1, min(ws.max_column, 80) + 1)
                    if normalise(ws.cell(r, c).value) in HEADER_ANCHORS
                )
                score = max(score, hits)
            if score > best_score:
                best, best_score = ws.title, score
        if best_score <= 0:
            raise InputSheetError(
                f"{self.path.name}: no sheet looks like a Listungen sheet "
                f"(no NART / Artikellangtext 1 header found)."
            )
        return best

    def _find_header_row(self, ws):
        for r in range(1, min(ws.max_row, 20) + 1):
            names = {normalise(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)}
            if all(a in names for a in HEADER_ANCHORS):
                return r
        raise InputSheetError(f"{self.path.name}: could not locate the header row.")

    def _map_columns(self, ws, header_row):
        cols, unmapped = {}, []
        for c in range(1, ws.max_column + 1):
            raw = ws.cell(header_row, c).value
            key = normalise(raw)
            if not key:
                continue
            canon = _SYNONYM_INDEX.get(key)
            if canon:
                cols.setdefault(canon, c)  # first occurrence wins (GTIN Stück before GTIN Stück2)
            else:
                unmapped.append(str(raw))
        return cols, unmapped

    # ------------------------------------------------------------------- read

    def rows(self):
        """Yield (excel_row, {canonical_name: value}) for every populated row.

        Values come from the cached-value workbook so formula columns such as
        the GTIN CONCATENATE resolve; formulas that Excel never evaluated come
        back as None and the caller falls back to deriving the value itself.
        """
        nart_col = self.columns["nart"]
        for r in range(self.header_row + 1, self.ws.max_row + 1):
            if self.ws_values.cell(r, nart_col).value in (None, ""):
                continue
            rec = {}
            for canon, c in self.columns.items():
                v = self.ws_values.cell(r, c).value
                if isinstance(v, str):
                    v = v.strip()
                    if v in ("", "-"):
                        v = None
                rec[canon] = v
            yield r, rec
