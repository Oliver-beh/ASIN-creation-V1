"""QA reporting.

Every run produces a verdict of UPLOAD READY or NOT UPLOAD READY, plus a
workbook that says exactly where each populated cell came from. The report is
the thing the person reviewing the upload actually reads, so it leads with the
blockers and keeps provenance one click away.

Blocking (NOT UPLOAD READY):
  * a PFLICHTFELD column left empty
  * a value rejected by the template's own dropdown
  * a duplicate GTIN or vendor SKU inside the batch
  * a GTIN that fails its GS1 check digit

Non-blocking (needs a human eye, upload can proceed):
  * a low-confidence derivation, currently only the scent fallback
  * a conditionally-required column left empty
  * GPSR compliance columns left empty
"""

from __future__ import annotations

import datetime as dt
from collections import Counter, defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from . import transforms as T

MAIN_BLUE = "002F40"
TEAL = "1CA0A5"
HIGHLIGHT = "F95954"
GREY = "F9F9F9"

GPSR_CODES = [
    "gpsr_safety_attestation#1.value",
    "gpsr_manufacturer_reference#1.gpsr_manufacturer_email_address",
    "dsa_responsible_party_address#1.value",
    "compliance_media#1.content_type",
]


class QAReport:
    def __init__(self, template, config, input_path, template_path):
        self.template = template
        self.config = config
        self.input_path = str(input_path)
        self.template_path = str(template_path)
        self.rows = []
        self.batch_issues = []
        self.notes = []

    def add_row(self, row_result):
        self.rows.append(row_result)

    # --------------------------------------------------------------- analysis

    def _duplicate_issues(self):
        issues = []
        for label, code in (("GTIN", "external_product_id#1.value"), ("Vendor SKU", "vendor_sku#1.value")):
            seen = defaultdict(list)
            for r in self.rows:
                f = r.fields.get(code)
                if f and f.value:
                    seen[str(f.value)].append(r.nart or str(r.source_row))
            for value, owners in seen.items():
                if len(owners) > 1:
                    issues.append(f"Duplicate {label} {value} on: {', '.join(owners)}")
        return issues

    def _checkdigit_issues(self):
        out = []
        for r in self.rows:
            f = r.fields.get("external_product_id#1.value")
            if f and f.value and not T.gtin_check_digit_ok(str(f.value)):
                out.append(f"GTIN {f.value} ({r.nart}) fails the GS1 check digit")
        return out

    def _missing_required(self):
        out = []
        required = set(self.template.required_field_codes())
        for r in self.rows:
            for code in required:
                f = r.fields.get(code)
                if f is None or f.value in (None, ""):
                    label = self.template.label(code)
                    out.append((r, code, label))
        return out

    def _rejected(self):
        return [(r, f) for r in self.rows for f in r.fields.values() if f.status == "rejected"]

    def _unresolved_conditional(self):
        out = []
        for r in self.rows:
            for f in r.fields.values():
                if f.status == "unresolved" and "BEDINGT" in (f.requirement or "").upper():
                    out.append((r, f))
        return out

    def _gpsr_gaps(self):
        present = [c for c in GPSR_CODES if self.template.resolve(c)]
        if not present:
            return []
        return [
            f"{len(present)} GPSR/compliance column(s) available in this template and left empty "
            f"({', '.join(present)}). Filling these at creation time avoids the PS01 / SDS rework loop."
        ]

    # ---------------------------------------------------------------- verdict

    def blockers(self):
        out = []
        out += [f"Required field empty: {label or code} (row {r.target_row}, {r.nart})"
                for r, code, label in self._missing_required()]
        out += [f"Value rejected by template dropdown: {f.label or f.code} (row {r.target_row}, {r.nart}) - {f.note}"
                for r, f in self._rejected()]
        out += self._duplicate_issues()
        out += self._checkdigit_issues()
        return out

    def warnings(self):
        out = []
        for r in self.rows:
            for f in r.reviews():
                out.append(f"Review: {f.label or f.code} = {f.value!r} (row {r.target_row}, {r.nart}) - {f.note}")
        for r, f in self._unresolved_conditional():
            out.append(f"Conditionally-required field empty: {f.label or f.code} (row {r.target_row}, {r.nart}) - {f.note}")
        out += self._gpsr_gaps()
        out += self.notes
        return out

    @property
    def status(self):
        return "UPLOAD READY" if not self.blockers() else "NOT UPLOAD READY"

    def summary_counts(self):
        prov = Counter()
        for r in self.rows:
            for f in r.fields.values():
                if f.status in ("ok", "review") and f.value not in (None, ""):
                    prov[f.provenance or "other"] += 1
        return prov

    # ----------------------------------------------------------------- output

    def write(self, path):
        wb = Workbook()
        self._sheet_summary(wb.active)
        self._sheet_blockers(wb.create_sheet("Blockers"))
        self._sheet_warnings(wb.create_sheet("Review"))
        self._sheet_provenance(wb.create_sheet("Field provenance"))
        wb.save(path)
        return path

    def _style_header(self, ws, headers, widths):
        fill = PatternFill("solid", fgColor=MAIN_BLUE)
        for i, (h, w) in enumerate(zip(headers, widths), start=1):
            c = ws.cell(1, i, h)
            c.font = Font(bold=True, color="FFFFFF", name="Lato")
            c.fill = fill
            ws.column_dimensions[c.column_letter].width = w
        ws.freeze_panes = "A2"

    def _sheet_summary(self, ws):
        ws.title = "Summary"
        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 78

        title = ws.cell(1, 1, "ASIN UPLOAD QA")
        title.font = Font(bold=True, size=16, color=MAIN_BLUE, name="Work Sans")
        period = ws.cell(1, 2, ".")
        period.font = Font(bold=True, size=16, color=HIGHLIGHT, name="Work Sans")
        ws.cell(1, 2).alignment = Alignment(horizontal="left")

        verdict = ws.cell(3, 1, "Status")
        verdict.font = Font(bold=True, name="Lato")
        v = ws.cell(3, 2, self.status)
        v.font = Font(bold=True, size=14, color="FFFFFF", name="Work Sans")
        v.fill = PatternFill("solid", fgColor=TEAL if self.status == "UPLOAD READY" else HIGHLIGHT)

        prov = self.summary_counts()
        rows = [
            ("Run at", dt.datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("Input sheet", self.input_path),
            ("Template", self.template_path),
            ("Product type", self.template.product_type),
            ("Locale", self.template.locale),
            ("Products written", len(self.rows)),
            ("Blocking issues", len(self.blockers())),
            ("Items to review", len(self.warnings())),
            ("", ""),
            ("Cells from a fixed value", prov.get("const", 0)),
            ("Cells copied from source", prov.get("source", 0)),
            ("Cells from a lookup table", prov.get("lookup", 0)),
            ("Cells from a rule", prov.get("derived", 0)),
        ]
        for i, (k, val) in enumerate(rows, start=5):
            a = ws.cell(i, 1, k)
            a.font = Font(bold=True, name="Lato", color=MAIN_BLUE)
            a.fill = PatternFill("solid", fgColor=GREY)
            ws.cell(i, 2, val).font = Font(name="Lato")

    def _sheet_blockers(self, ws):
        self._style_header(ws, ["#", "Blocking issue"], [6, 130])
        for i, b in enumerate(self.blockers(), start=1):
            ws.cell(i + 1, 1, i)
            ws.cell(i + 1, 2, b).font = Font(name="Lato")
        if not self.blockers():
            ws.cell(2, 2, "None. Every required column is populated and every value is a valid template option.")

    def _sheet_warnings(self, ws):
        self._style_header(ws, ["#", "Needs a human eye (not blocking)"], [6, 130])
        for i, w in enumerate(self.warnings(), start=1):
            ws.cell(i + 1, 1, i)
            ws.cell(i + 1, 2, w).font = Font(name="Lato")
        if not self.warnings():
            ws.cell(2, 2, "None.")

    def _sheet_provenance(self, ws):
        self._style_header(
            ws,
            ["Row", "NART", "Field code", "Column label", "Requirement", "Value", "Source", "Status", "Note"],
            [7, 18, 42, 32, 22, 46, 12, 12, 68],
        )
        r = 2
        for row in self.rows:
            for f in row.fields.values():
                if f.status == "absent":
                    continue
                ws.cell(r, 1, row.target_row)
                ws.cell(r, 2, row.nart)
                ws.cell(r, 3, f.code)
                ws.cell(r, 4, f.label)
                ws.cell(r, 5, f.requirement)
                ws.cell(r, 6, str(f.value) if f.value is not None else "")
                ws.cell(r, 7, f.provenance)
                sc = ws.cell(r, 8, f.status)
                if f.status in ("rejected", "unresolved"):
                    sc.font = Font(bold=True, color=HIGHLIGHT, name="Lato")
                elif f.status == "review":
                    sc.font = Font(bold=True, color=MAIN_BLUE, name="Lato")
                ws.cell(r, 9, f.note)
                r += 1
