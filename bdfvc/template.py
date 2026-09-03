"""Amazon Vendor Central bulk template reader / writer.

The template is the single source of truth. Nothing here knows anything about
Beiersdorf, cosmetics or any business rule - it only knows how Amazon builds
these workbooks:

  row 1  metadata blob (TemplateType, Locale, feedType, templateIdentifier ...)
  row 3  localised display label
  row 4  Amazon field code   <- everything is addressed by this, never by column letter
  row 5  requirement level   (PFLICHTFELD / BEDINGT ERFORDERLICH / OPTIONAL / EMPFOHLEN)
  row 6  Amazon's own example
  row 7+ data

Two kinds of dropdown exist and both are resolved here:

  * static      - a plain data validation pointing at a defined name
  * conditional - an x14 extension validation whose allowed values depend on the
                  value of one or more other cells in the same row

openpyxl silently drops the x14 block on save, so we read it straight out of the
xlsx zip and re-inject it after writing (see xlsm.py).
"""

from __future__ import annotations

import re
import warnings
import zipfile
from pathlib import Path

from openpyxl import load_workbook

warnings.filterwarnings("ignore")

TEMPLATE_SHEET_PREFIXES = ("Vorlage-", "Template-", "Modello-", "Plantilla-", "Modele-", "Modèle-")

LABEL_ROW = 3
FIELD_CODE_ROW = 4
REQUIREMENT_ROW = 5
EXAMPLE_ROW = 6
FIRST_DATA_ROW = 7

REQUIRED_TOKENS = ("PFLICHTFELD", "REQUIRED", "OBLIGATORIO", "OBBLIGATORIO")
CONDITIONAL_TOKENS = ("BEDINGT", "CONDITIONALLY")

# Sheet-name -> ordinal, so we can find the "Dropdown Lists" helper sheet in any locale.
DROPDOWN_SHEET_CANDIDATES = ("Dropdown Lists", "Dropdown-Listen", "Listes déroulantes")


def _col_letter_to_index(letters: str) -> int:
    n = 0
    for ch in letters:
        n = n * 26 + (ord(ch.upper()) - 64)
    return n


class ValidationSet:
    """The allowed values for one column, static or conditional."""

    __slots__ = ("static", "name_parts", "dep_cols")

    def __init__(self, static=None, name_parts=None, dep_cols=None):
        self.static = static  # list[str] | None
        self.name_parts = name_parts  # list[("lit", s) | ("dep", col_index)] | None
        self.dep_cols = dep_cols or []

    @property
    def is_conditional(self) -> bool:
        return self.name_parts is not None


class VCTemplate:
    def __init__(self, path):
        self.path = Path(path)
        self.wb = load_workbook(path, keep_vba=True)

        self.sheet_name = next(
            (s for s in self.wb.sheetnames if s.startswith(TEMPLATE_SHEET_PREFIXES)), None
        )
        if not self.sheet_name:
            raise ValueError(
                f"No product-template sheet in {self.path.name}. Sheets: {self.wb.sheetnames}"
            )
        self.ws = self.wb[self.sheet_name]
        self.product_type = self.sheet_name.split("-", 1)[1].strip()

        self.meta = self._parse_meta()
        self.locale = self.meta.get("Locale", "de_DE")

        self.field_codes = [self.ws.cell(FIELD_CODE_ROW, c).value for c in range(1, self.ws.max_column + 1)]
        self.labels = [self.ws.cell(LABEL_ROW, c).value for c in range(1, self.ws.max_column + 1)]
        self.requirements = [self.ws.cell(REQUIREMENT_ROW, c).value for c in range(1, self.ws.max_column + 1)]

        self.col_of = {}
        for i, fc in enumerate(self.field_codes):
            if fc:
                self.col_of.setdefault(str(fc).strip(), i + 1)

        self._lookup_table = self._load_dropdown_lookup()
        self.validations = self._load_validations()
        self._cascade_cache = {}

    # ---------------------------------------------------------------- metadata

    def _parse_meta(self) -> dict:
        meta = {}
        for cell in (c.value for c in self.ws[1] if c.value):
            for part in str(cell).split("&"):
                if "=" in part:
                    k, v = part.split("=", 1)
                    meta[k.strip()] = v.strip()
        return meta

    # ------------------------------------------------------------ field access

    def resolve(self, field_code: str):
        """Column index for a field code, tolerating the '#1.value' suffix variant."""
        fc = str(field_code).strip()
        if fc in self.col_of:
            return self.col_of[fc]
        alts = [f"{fc}#1.value"]
        if fc.endswith("#1.value"):
            alts.append(fc[: -len("#1.value")])
        for alt in alts:
            if alt in self.col_of:
                return self.col_of[alt]
        return None

    def requirement(self, field_code: str) -> str:
        col = self.resolve(field_code)
        return str(self.requirements[col - 1] or "") if col else ""

    def label(self, field_code: str) -> str:
        col = self.resolve(field_code)
        return str(self.labels[col - 1] or "") if col else ""

    def is_required(self, field_code: str) -> bool:
        return any(t in self.requirement(field_code).upper() for t in REQUIRED_TOKENS)

    def is_conditionally_required(self, field_code: str) -> bool:
        return any(t in self.requirement(field_code).upper() for t in CONDITIONAL_TOKENS)

    def required_field_codes(self):
        return [
            str(fc).strip()
            for fc, rq in zip(self.field_codes, self.requirements)
            if fc and rq and any(t in str(rq).upper() for t in REQUIRED_TOKENS)
        ]

    # --------------------------------------------------------------- dropdowns

    def _named_range_values(self, name: str):
        dn = self.wb.defined_names.get(name)
        if dn is None:
            return None
        out = []
        for sheet, ref in dn.destinations:
            try:
                cells = self.wb[sheet][ref.replace("$", "")]
            except Exception:
                continue
            if not isinstance(cells, tuple):
                cells = ((cells,),)
            elif cells and not isinstance(cells[0], tuple):
                cells = (cells,)
            for row in cells:
                for c in row:
                    if c.value is not None:
                        out.append(str(c.value))
        return out or None

    def _load_dropdown_lookup(self) -> dict:
        """Column A -> column B of the 'Dropdown Lists' sheet.

        Amazon uses this to turn a *displayed* value ('Körperpflege') into the
        token that forms part of a defined-name ('Koerperpflege' or similar).
        """
        sheet = next((s for s in self.wb.sheetnames if s in DROPDOWN_SHEET_CANDIDATES), None)
        if sheet is None:
            sheet = next((s for s in self.wb.sheetnames if "ropdown" in s or "istes" in s), None)
        if sheet is None:
            return {}
        ws = self.wb[sheet]
        out = {}
        for r in range(1, ws.max_row + 1):
            a, b = ws.cell(r, 1).value, ws.cell(r, 2).value
            if a is not None and b is not None:
                out.setdefault(str(a).strip(), str(b))
        return out

    def _parse_conditional_formula(self, formula: str):
        """Turn Amazon's cascade formula into (name_parts, dep_cols).

        The observed shape is always:
            INDIRECT("PT" & "a.value." & VLOOKUP(O7,'Dropdown Lists'!..) & ".b.value")
        We take the first INDIRECT(...) argument and read it as an ordered list of
        string literals and VLOOKUP-of-a-cell references.
        """
        idx = formula.find("INDIRECT(")
        if idx < 0:
            return None, []
        i = idx + len("INDIRECT(")
        depth = 1
        while i < len(formula) and depth:
            if formula[i] == "(":
                depth += 1
            elif formula[i] == ")":
                depth -= 1
            i += 1
        expr = formula[idx + len("INDIRECT(") : i - 1]

        parts, deps = [], []
        token_re = re.compile(r'"([^"]*)"|VLOOKUP\(\s*\$?([A-Z]{1,3})\$?\d+', re.I)
        for m in token_re.finditer(expr):
            if m.group(1) is not None:
                parts.append(("lit", m.group(1)))
            else:
                col = _col_letter_to_index(m.group(2))
                parts.append(("dep", col))
                deps.append(col)
        return (parts, deps) if parts else (None, [])

    def _load_validations(self) -> dict:
        vv = {}

        # --- static validations (openpyxl reads these fine) ------------------
        for dv in self.ws.data_validations.dataValidation:
            f = str(dv.formula1 or "")
            if not f:
                continue
            m = re.search(r'INDIRECT\("([^"]+)"\)', f)
            name = m.group(1) if m else f.lstrip("=")
            values = self._named_range_values(name)
            if not values:
                continue
            for rng in dv.sqref.ranges:
                for col in range(rng.min_col, rng.max_col + 1):
                    vv[col] = ValidationSet(static=values)

        # --- x14 conditional validations (read straight from the zip) --------
        for formula, sqref in self._raw_x14_validations():
            parts, deps = self._parse_conditional_formula(formula)
            if not parts:
                continue
            for ref in sqref.split():
                m = re.match(r"\$?([A-Z]{1,3})\$?\d+(?::\$?([A-Z]{1,3})\$?\d+)?", ref)
                if not m:
                    continue
                c0 = _col_letter_to_index(m.group(1))
                c1 = _col_letter_to_index(m.group(2)) if m.group(2) else c0
                for col in range(c0, c1 + 1):
                    vv[col] = ValidationSet(name_parts=parts, dep_cols=deps)
        return vv

    def _raw_x14_validations(self):
        """(formula, sqref) pairs from the worksheet's extLst block."""
        target = self._sheet_xml_path()
        if target is None:
            return []
        with zipfile.ZipFile(self.path) as z:
            xml = z.read(target).decode("utf-8", errors="ignore")
        idx = xml.find("<extLst")
        if idx < 0:
            return []
        ext = xml[idx:]
        out = []
        for m in re.finditer(r"<xm:f>(.*?)</xm:f>.*?<xm:sqref>(.*?)</xm:sqref>", ext, re.S):
            f = (
                m.group(1)
                .replace("&amp;", "&")
                .replace("&quot;", '"')
                .replace("&lt;", "<")
                .replace("&gt;", ">")
            )
            out.append((f, m.group(2)))
        return out

    def _sheet_xml_path(self):
        """Map the template sheet to its xl/worksheets/sheetN.xml entry."""
        with zipfile.ZipFile(self.path) as z:
            wb_xml = z.read("xl/workbook.xml").decode("utf-8", errors="ignore")
            rels = z.read("xl/_rels/workbook.xml.rels").decode("utf-8", errors="ignore")
        rid = None
        for m in re.finditer(r'<sheet[^>]*name="([^"]+)"[^>]*r:id="([^"]+)"', wb_xml):
            if m.group(1) == self.sheet_name:
                rid = m.group(2)
                break
        if rid is None:
            return None
        for m in re.finditer(r'<Relationship[^>]*Id="([^"]+)"[^>]*Target="([^"]+)"', rels):
            if m.group(1) == rid:
                t = m.group(2).lstrip("/")
                return t if t.startswith("xl/") else "xl/" + t
        return None

    # ----------------------------------------------------------- value lookup

    def allowed_values(self, col: int, row: int = None):
        """Allowed values for a column, resolving a cascade against the row if needed.

        Returns None when the column is free text.
        """
        vs = self.validations.get(col)
        if vs is None:
            return None
        if vs.static is not None:
            return vs.static
        if row is None:
            return None

        name = []
        for kind, payload in vs.name_parts:
            if kind == "lit":
                name.append(payload)
            else:
                dep_val = self.ws.cell(row, payload).value
                if dep_val in (None, ""):
                    return None  # driver not populated yet -> cannot resolve
                token = self._lookup_table.get(str(dep_val).strip())
                if token is None:
                    return None
                name.append(token)
        key = "".join(name)
        if key not in self._cascade_cache:
            self._cascade_cache[key] = self._named_range_values(key)
        return self._cascade_cache[key]

    def validate(self, col: int, value, row: int = None):
        """Return the canonically-cased allowed value, or None if not permitted.

        Free-text columns return the value unchanged. This never guesses: a value
        that is not literally in the template's own list is rejected.
        """
        allowed = self.allowed_values(col, row)
        if not allowed:
            return value if allowed is None else None
        want = str(value).strip()

        # Exact case first. Some lists carry several casings of the same word -
        # the brand column offers NIVEA, Nivea, NIVEA MEN, Nivea Men and
        # NIVEA Men - and the AVS guide requires the all-capitals spelling, so a
        # case-insensitive match alone would silently pick the wrong option.
        for opt in allowed:
            if str(opt).strip() == want:
                return opt
        folded = want.casefold()
        for opt in allowed:
            if str(opt).strip().casefold() == folded:
                return opt
        return None

    def cascade_depth(self, col: int) -> int:
        """0 for static/free columns, else 1 + depth of deepest driver."""
        vs = self.validations.get(col)
        if vs is None or not vs.is_conditional:
            return 0
        return 1 + max((self.cascade_depth(d) for d in vs.dep_cols), default=0)

    # ---------------------------------------------------------------- writing

    def clear_data_rows(self):
        for r in range(FIRST_DATA_ROW, self.ws.max_row + 1):
            for c in range(1, self.ws.max_column + 1):
                self.ws.cell(r, c).value = None

    def write(self, row: int, col: int, value):
        self.ws.cell(row, col).value = value

    def save(self, out_path):
        from .xlsm import save_preserving_extensions

        save_preserving_extensions(self.wb, self.path, out_path, self._sheet_xml_path())
