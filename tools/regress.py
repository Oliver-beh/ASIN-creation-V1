#!/usr/bin/env python3
"""Regenerate the known-good uploads and diff them cell by cell.

Four of the six verified examples were built on English templates, so an exact
string comparison would flag every enum. EQUIV maps the English value to its
German counterpart before comparing; anything left over is a real difference and
is either a bug here or a human error in the original.
"""

from __future__ import annotations

import datetime as dt
import os
import shutil
import sys
import tempfile
from collections import Counter
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import warnings

from openpyxl import load_workbook

from fill_bdf import build

warnings.filterwarnings("ignore")

# The verified fixtures (8 files: 3 blank templates, 3 Listungen, 5 human uploads).
# Colab mounted them at /mnt/user-data/uploads; point BDF_FIXTURES at wherever
# they live locally. Default is unchanged so existing invocations still work.
UP = Path(os.environ.get("BDF_FIXTURES", "/mnt/user-data/uploads"))
if not UP.is_dir():
    raise SystemExit(
        f"Fixture directory not found: {UP}\n"
        f"Set BDF_FIXTURES to the folder holding the verified templates, Listungen "
        f"and human uploads, e.g.\n  BDF_FIXTURES=~/bdf_fixtures python tools/regress.py"
    )

# English template value -> German equivalent
EQUIV = {
    "unit": "einheit", "yes": "ja", "no": "nein",
    "millimetres": "millimeter", "millimeter": "millimeter",
    "grams": "gramm", "gramm": "gramm", "kilograms": "kilogramm",
    "millilitres": "milliliter", "millilitre": "milliliter",
    "mass beauty": "massenschönheit", "multicolor": "mehrfarbig",
    "women": "damen", "men": "herren",
    "unisex adult": "unisex – erwachsene",
    "female": "weiblich", "male": "männlich",
    "germany": "deutschland", "poland": "polen", "spain": "spanien",
    "france": "frankreich", "italy": "italien", "netherlands": "niederlande",
    "turkey": "türkei", "china": "china",
    "transportation": "transport", "not applicable": "nicht zutreffend",
    "un regulatory id": "gefahrstoffkennung der vereinten nationen",
    "fresh": "frisch", "unscented": "unparfümiert", "citrus": "zitrus",
    "vanilla": "vanille", "rose": "rose",
    "normal": "normal", "sensitive": "empfindlich", "dry": "trocken", "all": "alle",
    "aerosol": "aerosol", "liquid": "flüssigkeit", "stick": "stift",
    "cream": "creme", "gel": "gel", "mousse": "mousse", "lotion": "lotion", "oil": "öl",
    "gtin": "gtin", "eur": "eur", "n/a": "n/a",
    "beiersdorf ag": "beiersdorf ag",
}

# Compared separately or not comparable across locales.
IGNORE = {
    "product_type#1.value",       # code vs display label differs by template build
    "rtip_vendor_code#1.value",   # full localised label differs by template build
    # The human copied the same dangerous-goods value into all five instances.
    # Instances 2-5 exist for *additional* regulations, so we write only #1.
    "supplier_declared_dg_hz_regulation#2.value",
    "supplier_declared_dg_hz_regulation#3.value",
    "supplier_declared_dg_hz_regulation#4.value",
    "supplier_declared_dg_hz_regulation#5.value",
}


def norm(v):
    if v is None or v == "":
        return None
    if isinstance(v, dt.datetime):
        return v.date().isoformat()
    if isinstance(v, dt.date):
        return v.isoformat()
    if isinstance(v, bool):
        return str(v)
    if isinstance(v, (int, float)):
        return f"{float(v):.4f}".rstrip("0").rstrip(".")
    s = str(v).strip()
    try:
        return f"{float(s.replace(',', '.')):.4f}".rstrip("0").rstrip(".")
    except ValueError:
        pass
    low = s.casefold()
    return EQUIV.get(low, low)


def read_rows(path, sheet=None):
    wb = load_workbook(path)
    sn = sheet or [s for s in wb.sheetnames if s.startswith(("Vorlage-", "Template-"))][0]
    ws = wb[sn]
    codes = {}
    for c in range(1, ws.max_column + 1):
        fc = ws.cell(4, c).value
        if fc:
            codes.setdefault(str(fc).strip(), c)
    out = {}
    sku_col = codes.get("vendor_sku#1.value")
    for r in range(7, ws.max_row + 1):
        sku = ws.cell(r, sku_col).value
        if not sku:
            continue
        out[str(sku).strip()] = {fc: ws.cell(r, c).value for fc, c in codes.items()}
    return out


CASES = [
    dict(
        name="Deodorants (30)",
        blank="Deodorants_2026-08-11T15_21_00Z.xlsm",
        expected="Deodorants_2026-04-09T10_51_00Z__2_.xlsm",
        input="Listungen_20260407.xlsx",
        filter="Body Deodorant",
    ),
    dict(
        name="Hair styling (18)",
        blank="Haarstyling-Mittel_2026-08-11T15_22_00Z.xlsm",
        expected="Hair_Styling_Agents_2026-04-09T10_48_00Z__1_.xlsm",
        input="Listungen_20260407.xlsx",
        filter="Hair Styling Agent",
    ),
    dict(
        name="Face care / Faceplus (4)",
        blank="Feuchtigkeitscreme_2026-07-24T08_39_00Z.xlsm",
        expected="Feuchtigkeitscreme_2026-07-24T08_39_00Z.xlsm",
        input="Listungen-Faceplus_20260722.xlsx",
        filter=None,
    ),
    dict(
        name="Feel Good Set, July rebuild (1)",
        blank="Hautpflegemittel_2026-08-11T15_20_00Z.xlsm",
        expected="Hautpflegemittel_2026-07-10T10_09_00Z_Feel_Good_Set.xlsm",
        input="Listungen-GP-20260526_Daten_erga_nzt_Feel_Good_Set_10-07-2026.xlsx",
        filter=None,
    ),
    dict(
        name="Gift sets, May batch (11)",
        blank="Hautpflegemittel_2026-08-11T15_20_00Z.xlsm",
        expected="Hautpflegemittel_2026-05-27T09_36_00Z_incl_cost_price.xlsm",
        input="Listungen-GP-20260526_Daten_erga_nzt_Feel_Good_Set_10-07-2026.xlsx",
        filter=None,
    ),
]


def run_case(case, verbose=False):
    tmp = Path(tempfile.mkdtemp())
    blank = tmp / "blank.xlsm"
    shutil.copy(UP / case["blank"], blank)
    report, out_xlsm, _ = build(
        blank, UP / case["input"], tmp / "out", filter_template=case["filter"]
    )
    got = read_rows(out_xlsm)
    want = read_rows(UP / case["expected"])

    common = set(got) & set(want)
    field_diffs = Counter()
    samples = {}
    matched = 0
    total = 0

    for sku in sorted(common):
        g, w = got[sku], want[sku]
        for fc, wv in w.items():
            if fc in IGNORE or norm(wv) is None:
                continue
            total += 1
            gv = g.get(fc)
            if norm(gv) == norm(wv):
                matched += 1
            else:
                field_diffs[fc] += 1
                samples.setdefault(fc, []).append((sku, wv, gv))

    pct = 100.0 * matched / total if total else 0.0
    print(f"\n{'=' * 78}")
    print(f"{case['name']}")
    print(f"{'=' * 78}")
    print(f"  rows generated {len(got):3}   overlap with expected {len(common):3}")
    print(f"  cells compared {total:4}   matched {matched:4}  ({pct:.1f}%)")
    print(f"  status         {report.status}   blockers {len(report.blockers())}")
    if field_diffs:
        print("  differing fields:")
        for fc, n in field_diffs.most_common():
            sku, wv, gv = samples[fc][0]
            print(f"    {n:3}x {fc}")
            print(f"          human {str(wv)[:56]!r}")
            print(f"          ours  {str(gv)[:56]!r}   (e.g. {sku})")
            if verbose:
                for sku, wv, gv in samples[fc][1:6]:
                    print(f"          .. {sku}: human {str(wv)[:34]!r} / ours {str(gv)[:34]!r}")
    return matched, total, len(report.blockers())


if __name__ == "__main__":
    verbose = "-v" in sys.argv
    tm = tt = tb = 0
    for case in CASES:
        m, t, b = run_case(case, verbose)
        tm += m
        tt += t
        tb += b
    print(f"\n{'=' * 78}")
    print(f"OVERALL  {tm}/{tt} cells match the verified human uploads ({100.0 * tm / tt:.1f}%)")
    print(f"         {tb} blocking issue(s) across all cases")
    print(f"{'=' * 78}\n")
